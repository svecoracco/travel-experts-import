"""Bouw het VAT Return-berekeningsworkbook met formules — poort van
`travel-experts-backend/apps/main/app/vat_return/excel_builder.py` (1-op-1,
zuivere openpyxl-code, geen Odoo-toegang).
"""

from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_vat_return_excel(
    period: str,
    start_data: Dict[str, Dict[str, float]],
    config: Dict[str, Any],
) -> bytes:
    """Bouw het VAT-return-berekeningsworkbook en retourneer als bytes.

    Het workbook bevat:
    1. Starttoestand-tabel
    2. Correcte-toestand-tabel (met Excel-formules)
    3. Correctie-entry-lijnen (met Excel-formules die naar bovenstaande tabellen verwijzen)
    """
    correction_mappings: List[Dict[str, str]] = config["correction_mappings"]
    remainder_grid: str = config["remainder_grid"]
    standard_vat_rate: float = float(config["standard_vat_rate"])
    correction_account: int = int(config["correction_account"])

    vat_codes = list(start_data.keys())

    # Bouw geordende grid-lijsten
    mapped_base_grids = [m["target_base_grid"] for m in correction_mappings]
    vat_grids = [m["source_vat_grid"] for m in correction_mappings]
    base_grids = [remainder_grid] + mapped_base_grids

    # Totaal aantal kolommen: VAT Code + base grids + vat grids
    num_base_cols = len(base_grids)
    num_vat_cols = len(vat_grids)
    total_data_cols = num_base_cols + num_vat_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "VAT Return"

    number_fmt = "#,##0.00"
    row = 1

    # ── Starttoestand ──────────────────────────────────────────

    # Rij 1: Sectieheader (merged)
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=1 + total_data_cols,
    )
    ws.cell(row=row, column=1, value="Start situation")
    ws.cell(row=row, column=1).font = _bold_font()
    row += 1

    # Rij 2: Groepheaders (merged)
    ws.cell(row=row, column=1, value="VAT Code")
    if num_base_cols > 0:
        ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=1 + num_base_cols,
        )
        ws.cell(row=row, column=2, value="Tax Grids - Base amounts")
    if num_vat_cols > 0:
        vat_start_col = 2 + num_base_cols
        ws.merge_cells(
            start_row=row,
            start_column=vat_start_col,
            end_row=row,
            end_column=vat_start_col + num_vat_cols - 1,
        )
        ws.cell(row=row, column=vat_start_col, value="Tax Grids - VAT Amounts")
    row += 1

    # Rij 3: Grid-kolom-subheaders
    ws.cell(row=row, column=1, value="VAT Code")
    col = 2
    for g in base_grids:
        ws.cell(row=row, column=col, value=g)
        col += 1
    for g in vat_grids:
        ws.cell(row=row, column=col, value=g)
        col += 1
    row += 1

    # Datarijen (startwaarden)
    start_data_first_row = row
    for code in vat_codes:
        ws.cell(row=row, column=1, value=code)
        col = 2
        for g in base_grids:
            cell = ws.cell(row=row, column=col, value=start_data[code].get(g, 0.0))
            cell.number_format = number_fmt
            col += 1
        for g in vat_grids:
            cell = ws.cell(row=row, column=col, value=start_data[code].get(g, 0.0))
            cell.number_format = number_fmt
            col += 1
        row += 1

    # Lege rij
    row += 1

    # ── Correcte toestand ────────────────────────────────────────

    # Sectieheader
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=1 + total_data_cols,
    )
    ws.cell(row=row, column=1, value="Correct situation")
    ws.cell(row=row, column=1).font = _bold_font()
    row += 1

    # Groepheaders (merged)
    ws.cell(row=row, column=1, value="VAT Code")
    if num_base_cols > 0:
        ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=1 + num_base_cols,
        )
        ws.cell(row=row, column=2, value="Tax Grids - Base amounts")
    if num_vat_cols > 0:
        vat_start_col = 2 + num_base_cols
        ws.merge_cells(
            start_row=row,
            start_column=vat_start_col,
            end_row=row,
            end_column=vat_start_col + num_vat_cols - 1,
        )
        ws.cell(row=row, column=vat_start_col, value="Tax Grids - VAT Amounts")
    row += 1

    # Grid-kolom-subheaders
    ws.cell(row=row, column=1, value="VAT Code")
    col = 2
    for g in base_grids:
        ws.cell(row=row, column=col, value=g)
        col += 1
    for g in vat_grids:
        ws.cell(row=row, column=col, value=g)
        col += 1
    row += 1

    # Datarijen met FORMULES
    corr_data_first_row = row

    # Bouw kolomindex-lookups (1-based kolomnummers)
    # Kolom 1 = VAT Code
    # Kolommen 2..2+num_base_cols-1 = base grids
    # Kolommen 2+num_base_cols..end = vat grids
    base_grid_cols = {}
    for i, g in enumerate(base_grids):
        base_grid_cols[g] = 2 + i

    vat_grid_cols = {}
    for i, g in enumerate(vat_grids):
        vat_grid_cols[g] = 2 + num_base_cols + i

    # Bouw mapping: source_vat_grid -> (target_base_grid, vat_col_letter)
    mapping_lookup = {}
    for m in correction_mappings:
        mapping_lookup[m["target_base_grid"]] = m["source_vat_grid"]

    for code_idx, code in enumerate(vat_codes):
        ws.cell(row=row, column=1, value=code)
        start_row_for_code = start_data_first_row + code_idx
        corr_row_for_code = row

        col = 2
        for g in base_grids:
            if g == remainder_grid:
                # Remainder = som(start mapped bases) - som(gecorrigeerde mapped bases)
                start_refs = []
                corr_refs = []
                for bg in mapped_base_grids:
                    start_col_letter = get_column_letter(base_grid_cols[bg])
                    corr_col_letter = get_column_letter(base_grid_cols[bg])
                    start_refs.append(f"{start_col_letter}{start_row_for_code}")
                    corr_refs.append(f"{corr_col_letter}{corr_row_for_code}")
                start_sum = "+".join(start_refs)
                corr_sum = "+".join(corr_refs)
                formula = f"=({start_sum})-({corr_sum})"
                cell = ws.cell(row=row, column=col, value=formula)
            elif g in mapping_lookup:
                # Gecorrigeerde base = VAT-bedrag / tarief
                source_vat = mapping_lookup[g]
                vat_col_letter = get_column_letter(vat_grid_cols[source_vat])
                formula = f"={vat_col_letter}{corr_row_for_code}/{standard_vat_rate}"
                cell = ws.cell(row=row, column=col, value=formula)
            else:
                cell = ws.cell(row=row, column=col, value=start_data[code].get(g, 0.0))
            cell.number_format = number_fmt
            col += 1

        # VAT-grids: verwijs naar startwaarden (ongewijzigd)
        for g in vat_grids:
            vat_col_letter = get_column_letter(vat_grid_cols[g])
            formula = f"={vat_col_letter}{start_row_for_code}"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = number_fmt
            col += 1

        row += 1

    # Lege rij
    row += 1

    # ── Correction Entry ─────────────────────────────────────────

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=6,
    )
    ws.cell(row=row, column=1, value="Correction Entry")
    ws.cell(row=row, column=1).font = _bold_font()
    row += 1

    # Headers: Account | Description | (leeg) | (leeg) | Tax grid | Amount
    headers = ["Account", "Description", "", "", "Tax grid", "Amount"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    row += 1

    # Correctielijnen met formules
    for code_idx, code in enumerate(vat_codes):
        start_row_for_code = start_data_first_row + code_idx
        corr_row_for_code = corr_data_first_row + code_idx

        for bg in mapped_base_grids:
            base_col_letter = get_column_letter(base_grid_cols[bg])

            # Reversal-lijn: negeer startwaarde
            ws.cell(row=row, column=1, value=correction_account)
            ws.cell(
                row=row,
                column=2,
                value=f"Correction VAT {period} - {code} - Tax grid: {bg} - Start situation",
            )
            ws.cell(row=row, column=5, value=bg)
            cell = ws.cell(
                row=row,
                column=6,
                value=f"=-{base_col_letter}{start_row_for_code}",
            )
            cell.number_format = number_fmt
            row += 1

            # Gecorrigeerde lijn: verwijs naar gecorrigeerde waarde
            ws.cell(row=row, column=1, value=correction_account)
            ws.cell(
                row=row,
                column=2,
                value=f"Correction VAT {period} - {code} - Tax grid: {bg} - Corrected situation",
            )
            ws.cell(row=row, column=5, value=bg)
            cell = ws.cell(
                row=row,
                column=6,
                value=f"={base_col_letter}{corr_row_for_code}",
            )
            cell.number_format = number_fmt
            row += 1

        # Remainder-lijn
        remainder_col_letter = get_column_letter(base_grid_cols[remainder_grid])
        ws.cell(row=row, column=1, value=correction_account)
        ws.cell(
            row=row,
            column=2,
            value=f"Correction VAT {period} - {code} - Tax grid: {remainder_grid} - Corrected situation",
        )
        ws.cell(row=row, column=5, value=remainder_grid)
        cell = ws.cell(
            row=row,
            column=6,
            value=f"={remainder_col_letter}{corr_row_for_code}",
        )
        cell.number_format = number_fmt
        row += 1

    # Autofit kolombreedtes (benaderend)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bold_font():
    """Retourneer een vetgedrukt openpyxl Font."""
    from openpyxl.styles import Font

    return Font(bold=True)
