"""Commission Excel reader — poort van
`travel-experts-backend/apps/main/app/plugins/commission/excel_reader.py`
(1-op-1, zuivere openpyxl-code, geen Odoo-toegang).

Leest een commission billing Excel-bestand (bv. van BFE / ARO) met:
- Metadata-zone (variabel aantal rijen): factuurmetadata in merged-cell label/waarde-paren
- Titelrij (net vóór de kolomheaders): overslaan
- Kolomheaders-rij (dynamisch — gedetecteerd door te scannen op bekende kolomnamen)
- Datarijen (header_row + 1 en verder): één rij per dossier
- Gele totaalrij: markeert het einde van de data (alles eronder genegeerd)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import Cell

logger = logging.getLogger(__name__)

YELLOW_RGB = "FFFFFF00"

# Bekende kolomheadernamen om de headerrij te detecteren.
# We zoeken naar "Dossier" (verplicht) plus minstens één andere.
_KNOWN_HEADERS = {
    "agent",
    "dossier",
    "date de départ",
    "date de depart",
    "nom client",
    "net facturé",
    "net facture",
    "achats",
    "comm. payée bts",
    "comm. payee bts",
    "coût cc",
    "cout cc",
    "solde (facturé)",
    "solde (facture)",
    "commission te",
    "marge (facturé) %",
    "marge (facture) %",
    "dû",
    "du",
    "comm. à payer",
    "comm. a payer",
}

# Metadatalabels (kolom A) → veldnaam-mapping
_META_LABELS: Dict[str, str] = {
    "leverancier": "supplier_name",
    "datum factuur": "invoice_date",
    "totaal bedrag": "total_amount",
    "opmerking - factuurnummer": "invoice_ref",
    "mededeling": "narration",
    "btw code": "vat_code",
}


@dataclass
class CommissionHeaderData:
    supplier_name: Optional[str]
    invoice_date: Optional[date]
    total_amount: Optional[float]
    invoice_ref: Optional[str]
    narration: Optional[str]
    vat_code: Optional[str]


@dataclass
class CommissionRow:
    dossier: str  # "Dossier"-kolom
    client_name: str  # "Nom client"-kolom
    amount: float  # "Comm. à payer"-kolom (signed: + = kost, - = terugbetaling)


@dataclass
class CommissionFileData:
    header: CommissionHeaderData
    rows: List[CommissionRow]
    header_row: int  # 1-based rijnummer van de kolomheaders


def _get_fill_rgb(cell: Cell) -> Optional[str]:
    """Retourneer de fill-foreground-RGB-string voor een cel, of None."""
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb and rgb != "00000000":
                return rgb
    except Exception:
        pass
    return None


def _cell_value(ws: Any, row: int, col: int) -> Any:
    """Haal celwaarde op."""
    return ws.cell(row=row, column=col).value


def _first_nonempty_after_merge(
    ws: Any, row: int, start_col: int = 2, max_col: int = 14
) -> Any:
    """Zoek de eerste niet-lege celwaarde in een rij vanaf start_col."""
    for col in range(start_col, max_col + 1):
        val = _cell_value(ws, row, col)
        if val is not None and str(val).strip() != "":
            return val
    return None


def _parse_date_value(val: Any) -> Optional[date]:
    """Parse een waarde naar een date-object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val: Any) -> Optional[float]:
    """Parse een waarde naar float, met Europese komma-decimalen."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _normalize(text: str) -> str:
    """Normaliseer een header-string voor vergelijking (lowercase, strip whitespace/newlines)."""
    return " ".join(text.lower().split())


def _find_header_row(ws: Any, max_scan: int = 30) -> Optional[int]:
    """Scan rijen om de kolomheaders-rij te vinden.

    Retourneert het 1-based rijnummer, of None indien niet gevonden.
    De headerrij wordt herkend doordat hij "Dossier" en minstens twee andere
    bekende kolomnamen bevat.
    """
    for row_idx in range(1, min(ws.max_row + 1, max_scan + 1)):
        found_names: List[str] = []
        has_dossier = False
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            val = _cell_value(ws, row_idx, col_idx)
            if val is None:
                continue
            norm = _normalize(str(val))
            if norm in _KNOWN_HEADERS:
                found_names.append(norm)
            if "dossier" in norm:
                has_dossier = True
        # Vereis "Dossier" + minstens 2 andere bekende headers
        if has_dossier and len(found_names) >= 3:
            logger.info(
                "[commission] Header row found at row %d (matched: %s)",
                row_idx,
                found_names,
            )
            return row_idx
    return None


def _build_column_map(ws: Any, header_row: int) -> Dict[str, int]:
    """Bouw een mapping van genormaliseerde headernaam → 1-based kolomindex."""
    col_map: Dict[str, int] = {}
    for col_idx in range(1, min(ws.max_column + 1, 20)):
        val = _cell_value(ws, header_row, col_idx)
        if val is not None:
            norm = _normalize(str(val))
            col_map[norm] = col_idx
    return col_map


def _resolve_col(col_map: Dict[str, int], *candidates: str) -> Optional[int]:
    """Zoek de kolomindex voor de eerste matchende kandidaatnaam."""
    for name in candidates:
        if name in col_map:
            return col_map[name]
    return None


def _parse_metadata(ws: Any, header_row: int) -> CommissionHeaderData:
    """Extraheer metadata uit rijen boven de headerrij.

    Scant rijen 1 .. header_row-1 op bekende labels in kolom A, en leest
    dan de waarde uit de eerste niet-lege cel rechts ervan.
    """
    meta: Dict[str, Any] = {}
    for row_idx in range(1, header_row):
        label_val = _cell_value(ws, row_idx, 1)
        if label_val is None:
            continue
        label_norm = _normalize(str(label_val))
        field_name = _META_LABELS.get(label_norm)
        if field_name:
            meta[field_name] = _first_nonempty_after_merge(ws, row_idx)

    supplier = meta.get("supplier_name")
    invoice_ref = meta.get("invoice_ref")
    narration = meta.get("narration")
    vat_code = meta.get("vat_code")

    return CommissionHeaderData(
        supplier_name=str(supplier).strip() if supplier is not None else None,
        invoice_date=_parse_date_value(meta.get("invoice_date")),
        total_amount=_parse_float(meta.get("total_amount")),
        invoice_ref=str(invoice_ref).strip() if invoice_ref is not None else None,
        narration=str(narration).strip() if narration is not None else None,
        vat_code=str(vat_code).strip() if vat_code is not None else None,
    )


def _is_yellow_row(ws: Any, row: int) -> bool:
    """Controleer of kolom A van de gegeven rij een gele fill heeft."""
    rgb = _get_fill_rgb(ws.cell(row=row, column=1))
    return rgb == YELLOW_RGB


def _parse_data_rows(
    ws: Any,
    data_start_row: int,
    dossier_col: int,
    client_col: int,
    amount_col: int,
) -> List[CommissionRow]:
    """Parse datarijen vanaf data_start_row, stop bij gele totaalrij of lege rij."""
    rows: List[CommissionRow] = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        # Stop bij gele totaalrij
        if _is_yellow_row(ws, row_idx):
            logger.info(
                "[commission] Yellow total row found at row %d — stopping", row_idx
            )
            break

        dossier_val = _cell_value(ws, row_idx, dossier_col)
        client_val = _cell_value(ws, row_idx, client_col)
        amount_val = _cell_value(ws, row_idx, amount_col)

        dossier = str(dossier_val).strip() if dossier_val is not None else ""
        amount = _parse_float(amount_val)

        # Stop bij volledig lege rij (geen dossier EN geen bedrag)
        if not dossier and amount is None:
            logger.info("[commission] Empty row at row %d — stopping", row_idx)
            break

        client_name = str(client_val).strip() if client_val is not None else ""

        rows.append(
            CommissionRow(
                dossier=dossier,
                client_name=client_name,
                amount=amount if amount is not None else 0.0,
            )
        )

    return rows


def _find_columns_and_parse(ws: Any) -> Tuple[int, int, int, int, int]:
    """Zoek de headerrij en resolve de drie kern-kolomindices.

    Retourneert (header_row, data_start_row, dossier_col, client_col, amount_col).
    Werpt ValueError als de headerrij of verplichte kolommen niet gevonden worden.
    """
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError(
            "Cannot find the column header row. Expected a row containing "
            "'Agent', 'Dossier', 'Nom client', 'Comm. à payer', etc."
        )

    col_map = _build_column_map(ws, header_row)

    dossier_col = _resolve_col(col_map, "dossier")
    client_col = _resolve_col(col_map, "nom client")
    amount_col = _resolve_col(col_map, "comm. à payer", "comm. a payer")

    if dossier_col is None:
        raise ValueError("Cannot find 'Dossier' column in header row")
    if client_col is None:
        raise ValueError("Cannot find 'Nom client' column in header row")
    if amount_col is None:
        raise ValueError("Cannot find 'Comm. à payer' column in header row")

    data_start_row = header_row + 1
    return header_row, data_start_row, dossier_col, client_col, amount_col


def read_commission_excel(path: str) -> CommissionFileData:
    """Lees een commission billing Excel-bestand en retourneer geparste data."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row, data_start, dossier_col, client_col, amount_col = (
        _find_columns_and_parse(ws)
    )

    header = _parse_metadata(ws, header_row)
    rows = _parse_data_rows(ws, data_start, dossier_col, client_col, amount_col)

    logger.info(
        "[commission] read_commission_excel: %d data rows from %s "
        "(header_row=%d supplier=%s invoice_ref=%s total=%s)",
        len(rows),
        path,
        header_row,
        header.supplier_name,
        header.invoice_ref,
        header.total_amount,
    )
    return CommissionFileData(header=header, rows=rows, header_row=header_row)


def validate_commission_columns(path: str) -> List[str]:
    """Valideer de structuur van het commission Excel-bestand.

    Retourneert een lijst met fouten (leeg = geldig).
    """
    errors: List[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as exc:
        return [str(exc)]

    # Zoek de headerrij dynamisch
    header_row = _find_header_row(ws)
    if header_row is None:
        errors.append(
            "Cannot find the column header row. Expected a row containing "
            "'Dossier', 'Nom client', 'Comm. à payer', etc."
        )
        return errors

    # Verifieer dat verplichte kolommen bestaan
    col_map = _build_column_map(ws, header_row)
    if _resolve_col(col_map, "dossier") is None:
        errors.append("Missing required column 'Dossier'")
    if _resolve_col(col_map, "nom client") is None:
        errors.append("Missing required column 'Nom client'")
    if _resolve_col(col_map, "comm. à payer", "comm. a payer") is None:
        errors.append("Missing required column 'Comm. à payer'")

    # Check dat de metadata-zone minstens "Leverancier" ergens boven de headerrij heeft
    found_leverancier = False
    for row_idx in range(1, header_row):
        val = _cell_value(ws, row_idx, 1)
        if val is not None and "leverancier" in _normalize(str(val)):
            found_leverancier = True
            break
    if not found_leverancier:
        errors.append(
            "Cannot find 'Leverancier' label in the metadata zone "
            "above the column headers"
        )

    return errors


def count_commission_data_rows(path: str) -> int:
    """Tel datarijen (tussen kolomheaders en gele totaalrij)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header_row, data_start, dossier_col, client_col, amount_col = (
            _find_columns_and_parse(ws)
        )
        rows = _parse_data_rows(ws, data_start, dossier_col, client_col, amount_col)
        return len(rows)
    except Exception:
        return 0
