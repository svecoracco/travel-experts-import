"""Eenmalig hulpscript om de minimale synthetische parity-harness-fixtures te
genereren onder `functions/tests/fixtures/` (fase 2.4 — "waar een echte
fixture ontbreekt voor een plugin: bouw een minimale synthetische fixture").

Niet onderdeel van de pytest-collectie (`generate_` prefix, geen `test_`).
Draai handmatig: `python functions/tests/fixtures/generate_fixtures.py`
(al gedraaid; de resulterende bestanden staan gecommit in deze map).

**Bewuste afwijking t.o.v. de fase-2-opdracht (dataveiligheid)**: voor
`vivawallet` was een ECHTE sample-fixture beschikbaar
(`C:\\github\\travel-experts\\files\\VIVA WALLET 01.03 - 05.03.2026.xlsx`,
zoals expliciet gesuggereerd in de opdracht). Bij inspectie bleek dat
bestand echte klant-PII te bevatten (kolommen "E-mail", "Phone",
"Card Holder", "Card Number") — ongeschikt om als gecommitte test-fixture in
de nieuwe repo te zetten. Er is daarom **niet** met dat bestand gewerkt; in
plaats daarvan is ook voor `vivawallet` een minimale synthetische fixture
gebouwd (zelfde patroon als de andere 7 plugins hieronder), met uitsluitend
de kolommen die `plugins/vivawallet/excel_reader.py::REQUIRED_COLS`/
`TEXT_COLS` daadwerkelijk gebruiken, gevuld met evident fictieve waarden.
Zie het fase-2-eindrapport voor deze afweging.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

FIXTURES_DIR = Path(__file__).parent


def _write_airplus() -> None:
    df = pd.DataFrame(
        {
            "Supplier code": ["SUP001"],
            "Invoice number": ["INV-0001"],
            "Reference": ["Trip to Paris"],
            "File number": ["FN-100"],
            "Currency": ["EUR"],
            "VAT code": [""],
            "Total amount": [123.45],
            "Bookings date": ["2026-02-15"],
            "Ledger account": [604000],
        }
    )
    df.to_excel(FIXTURES_DIR / "airplus_sample.xlsx", index=False)


def _write_divers() -> None:
    df = pd.DataFrame(
        {
            "Reference": ["Divers ref 1"],
            "Invoice date": ["2026-02-10"],
            "Invoice number": ["DIV-0001"],
            "net amount": [50.0],
            "ledger account": [612000],
            "Vat code": [""],
            "File number": ["FN-200"],
            "Supplier code": ["SUP002"],
            "Traveller": ["J. Doe"],
        }
    )
    df.to_excel(FIXTURES_DIR / "divers_sample.xlsx", index=False)


def _write_rail() -> None:
    df = pd.DataFrame(
        {
            "ISSUE_ID": ["1234567"],
            "PNR": ["ABCDEF"],
            "DNR_ID": ["9876543"],
            "TICKET_NBR": ["1234567"],
            "TRANSACTION_PRICE": ["43,00"],
            "NET_AMOUNT": [43.0],
            "OFFICIAL_DOC_NUMBER": ["ODN-0001"],
            "DEPARTURE_DATE": ["15.02.2026"],
        }
    )
    df.to_excel(FIXTURES_DIR / "rail_sample.xlsx", index=False)


def _write_bsp() -> None:
    header = (
        "BSP;Airline Code;Airline Name;Document Type;Document Number;Issue Date;"
        "Period;Currency;Cash;Payment Card;EasyPay;Net to be Paid;Comments\n"
    )
    row = (
        "220;074;KLM;TKTT;1234567890;2026-02-15;2026021;EUR;0.00;100.00;0.00;0.00;"
        'CC VI 492019XXXXXX2533 CA\n'
    )
    (FIXTURES_DIR / "bsp_sample.csv").write_text(header + row, encoding="utf-8-sig")


def _write_ibanfirst() -> None:
    header = (
        "Identificatie,Boekingdatum,Valutadatum,Toestand,Type,Tegenpartij,"
        "Beschrijving,Communicatie,Interne referentie,Gedebiteerd bedrag,"
        "Gecrediteerd bedrag,Valuta\n"
    )
    row = (
        "TX-0001,15/02/2026,15/02/2026,Definitief,Payment,Some Counterparty,"
        "Invoice payment,COMM-1,INT-1,0,150.00,EUR\n"
    )
    (FIXTURES_DIR / "ibanfirst_sample.csv").write_text(header + row, encoding="utf-8")


def _write_commission() -> None:
    # Metadata-zone (rijen 1-6) + headerrij (rij 7) + één datarij (rij 8) +
    # gele totaalrij (rij 9) — zie plugins/commission/excel_reader.py.
    wb = Workbook()
    ws = wb.active
    ws.title = "Commission"

    meta_rows = [
        ("Leverancier", "Commission Supplier NV"),
        ("Datum factuur", "15/02/2026"),
        ("Totaal bedrag", 100.0),
        ("Opmerking - factuurnummer", "COMM-INV-0001"),
        ("Mededeling", "Commission narration"),
        ("Btw code", ""),
    ]
    for i, (label, value) in enumerate(meta_rows, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    header_row = len(meta_rows) + 1  # rij 7
    headers = ["Agent", "Dossier", "Date de départ", "Nom client", "Comm. à payer"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)

    data_row = header_row + 1  # rij 8
    ws.cell(row=data_row, column=1, value="Agent1")
    ws.cell(row=data_row, column=2, value="FN-300")
    ws.cell(row=data_row, column=3, value="20/02/2026")
    ws.cell(row=data_row, column=4, value="Client X")
    ws.cell(row=data_row, column=5, value=100.0)

    total_row = data_row + 1  # rij 9, gele totaalrij (stopt de dataparsing)
    yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    ws.cell(row=total_row, column=1, value="Total").fill = yellow

    wb.save(FIXTURES_DIR / "commission_sample.xlsx")


def _write_vivawallet() -> None:
    # Uitsluitend de kolommen die plugins/vivawallet/excel_reader.py leest
    # (TEXT_COLS + REQUIRED_COLS + de optionele Date/Time/*Description-velden
    # die transform.py gebruikt) — geen PII-kolommen (geen E-mail/Phone/
    # Card Holder/Card Number, zie de kop-comment hierboven).
    df = pd.DataFrame(
        {
            "Order Code": ["ORD-0001", "ORD-0002"],
            "Merchant Reference": ["8400165", ""],
            "Transaction ID": ["TXN-0001", "TXN-0002"],
            "Card Type": ["VISA", "MASTERCARD"],
            "Date": ["2026-02-15", "2026-02-16"],
            "Time": ["10:00:00", "11:30:00"],
            "Source Description": ["Card payment", "Card payment"],
            "Customer Description": ["Synthetic test transaction 1", "Synthetic test transaction 2"],
            "Amount": [120.0, 45.5],
            "NetAmount": [118.0, 44.0],
            "Clearance Date": ["2026-02-15", "2026-02-16"],
        }
    )
    df.to_excel(FIXTURES_DIR / "vivawallet_sample.xlsx", index=False, engine="openpyxl")


def _write_tui() -> None:
    # Bestandsnaam-masker DOM_JET_{group}_{datum}_{ref}_...
    # 16 '#'-gescheiden kolommen (col_ignored_1 .. col_ignored_7), geen header
    # (.csv-variant) — zie plugins/tui/csv_reader.py::COLUMN_NAMES.
    fields = [
        "1201",  # col_ignored_1
        "GRP1",  # col_group
        "",  # col_comm_ref
        "",  # col_date1 (unused)
        "",  # col_date2 (unused)
        "005678",  # col_ticket_ref
        "20260215",  # col_departure
        "Ticket description",  # col_description
        "EUR",  # col_currency
        "",  # col_ignored_2
        "123.45",  # col_amount
        "",  # col_ignored_3
        "",  # col_ignored_4
        "",  # col_ignored_5
        "",  # col_ignored_6
        "",  # col_ignored_7
    ]
    row = "#".join(fields) + "\n"
    (FIXTURES_DIR / "DOM_JET_GRP1_20260215_REF001_sample.csv").write_text(
        row, encoding="utf-8"
    )


def main() -> None:
    _write_airplus()
    _write_divers()
    _write_rail()
    _write_bsp()
    _write_ibanfirst()
    _write_commission()
    _write_tui()
    _write_vivawallet()
    print("Synthetic fixtures written to", FIXTURES_DIR)


if __name__ == "__main__":
    main()
